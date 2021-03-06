#For Broker dealers don't collect SIC code and change openpyxl code to account for missing column.
#List of filings to look for, any filing could be added
# Operating Companies: 
Filing_list = ['10-K', '20-F', '10KSB', 'S-1','40-F', 'S-4', '1-F', 'REVOKED']
#Broker Dealer filings 
# Filing_list = ['X-17A-5', 'FOCUSN']

import time


import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
from pandas import ExcelWriter
import re
import os
import string

#User enters the file location information
s = raw_input('Input the file location ->')
s2 = raw_input('Input the file name ->')
s3 = raw_input('Input the save file name ->')
#Start a timer for how long it takes to run the code
start_time = time.time()
#For the main file path, one needs to make sure they are forward slashes
s = string.replace(s, "\\", '/')

# Combining the main folder location and the xlsx file that has the list of ciks
main_location = s + '/' + s2
# Combining the temporary csv files for the different results with the folder location
result_company_info = s + '/result_company_info.csv'
result_filing = s + '/result_filing.csv'
result_wrong = s + '/result_wrong.csv'
results = s + '/' + s3





#Function for saving exsisting and non exsisting files
def savingout(save,dataframe):
	if not os.path.isfile(save):
		dataframe.to_csv(save, index=False)
	else: 
		dataframe.to_csv(save,mode = 'a',header=False, index=False)

#Function for looking through lines of html, fininding the matching regular expression, and then returing the matching line
def filterPick(lines, regex):
    matches = map(re.compile(regex).search, lines)
    return [m.group(1) for m in matches if m]

	
#Pandas reading in the main list of ciks
in_file = pd.read_excel(main_location)
cik = in_file['cik'].tolist()

#Empty dataframes where different reuslts will be stored
wrong_df = pd.DataFrame()

results_df2 = pd.DataFrame()
results2_df = pd.DataFrame()

# Loop keeps track of how many 500 iterations have passed
# Time2 keeps track of how many iterations have happenned

loop = 1
time2 = 0
with requests.Session() as s:
    s.get('https://www.sec.gov/')
	
#Start of the loop, i is the cik for each row in the 'cik' list
for i in cik:
	time2 = time2 + 1
	#Main sec header which never changes
	header = 'https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK='
	
	#Main sec footer can be changed to show different information, but here we are looking in first 100 filings for a given cik
	footer = '&type=&dateb=&owner=exclude&count=100'
	#create url to parse through sec header + cik + fotter
	url = header + str(i) + footer
	#r here is the url that we quieried
	r = s.get(url)
	#soup is an html tree where we took the content of r and put it in the html tree. IF you look up a given link, it should be the same html tree that you would see on that page
	soup = bs(r.content,'html.parser')
	#First think is to check if the cik is a real company, thus we search the soup for the phase
	no_cik = soup.find(text=re.compile("No matching CIK"))
	#If we find the phrase in the soup we save that cik as wrong in our results and start the loop over at the next cik
	if no_cik == u'No matching CIK.':
		#no_cik is in utf format so we translate it to ascii format in order to save it to a csv
		wrong = no_cik.encode('ascii','ignore')
		#create the columns for a dataframe
		columns = {'Wrong' : wrong, 'CIK': i}
		#create a dataframe with the given columns
		df = pd.DataFrame(columns, index =[0])
		#append the columns to the dataframe storing wrong results
		wrong_df = wrong_df.append(df)
		#Tells the use the cik doesn't exsist
		print "\n Whoomps there it is %s" %(i) 
		progress = round((float(time2) / len(cik)) *100)
		#Tracking progress
		print "\n %d Percent Done" % (progress)
		#Following if the results should be save out yet or no.
		if time2 >= 500 * loop:
			print "Saving Results"
			loop = 1 + loop
			savingout(result_company_info, results2_df)
			savingout(result_filing, results_df2)
			savingout(result_wrong, wrong_df)
			wrong_df = pd.DataFrame()
			results_df2 = pd.DataFrame()
			results2_df = pd.DataFrame()
			continue
		else:
			continue
	# A good portion of the code uses try/except loops. In this case, if the company does not have a company info table then we say 'No Descritption'
	try:
		#There are a lot of ways to use soup.find(), here we try to find the item in the html tree that contains the class for company name. Then store the results as p
		p = soup.find("span", { "class" : "companyName" })
		
		#Blank Dictionary for Company Names
		Name = []
		#p refers to the result we found where there was a company name. THe only item in it should be the company name and we append that to the empty dictionary
		for x in p:
			Name.append(x)
		# This is too keep the first Name we append in the column. I forgot why I did this to be honest.
		Name = Name[0]
		# Now we want to scrape the company information. This part is kinda tricky. We grab everything under a tag of p and has the class of identinfo.
		p2 = soup.find_all("p", { "class" : "identInfo" })[0].find_all_next(string = True)
		# The next part takes each of the lines we found  in the tree and puts it into a dictionary that is easier to navigate and scrape info from. 
		#Using the regular expression function written in the beginning, we scrape the line for the corresponding information we want. Luckily the SEC has consistent format across each page and I have yet to encouter an error doing this.
		body_l = len(soup.find('p'))
		Company_Info = []
		
		for cmpin in p2[:body_l]:
			if  len(str(cmpin).strip()) > 1:
				y = str(cmpin).strip()
				y = y.encode('ascii','ignore')
				Company_Info.append(y)
		
		State_Location = []
		State_Incorporation = []
		Fiscal_Year_End = []
		SIC_Code = []
		#SIC Code
		if filterPick(Company_Info, '(SIC)'):
			SIC_Code = '-'.join(Company_Info[0:3])
		else:
			pass
			
		if filterPick(Company_Info, '(State)'):
			for q, j in enumerate(Company_Info):
				#State where the HQ is located
				if re.search('(location)',j):
					w = q+1
					ww = Company_Info[w]
					State_Location.append(ww)
				#State where they are incorporated
				elif re.search('(of Inc)',j):
					w = q+1
					ww = Company_Info[w]
					State_Incorporation.append(ww)
				else:
					pass
		else:
			pass
		#Current fiscal year end
		if filterPick(Company_Info, '(Fiscal)' ):
			for q, j in enumerate(Company_Info):
				if re.search('(Fiscal)',j):
					ww = Company_Info[q]
					Fiscal_Year_End.append(ww)	
		else:
			pass

		Other_Name = []
		#Scrape the other former company names
		if filterPick(Company_Info, '(formerly)'):
			for x in Company_Info:
				if re.search('(formerly)',x):
					x = x.encode('ascii','ignore')
					Other_Name.append(x)
		else:
			pass
		Other_Name = ''.join(Other_Name)
		
		#Put the company information in columns for a pandas dictionary
		columns2 = {'CIK': i, 'Company_Name' :Name, 'SIC_Code' : SIC_Code, 'State_Location' : State_Location, 'Fiscal_Year_End': Fiscal_Year_End, 'Other_Name' : Other_Name, 'State_Incorporation' : State_Incorporation, 'Link' :url}
		
		#Deletes empty columns if no information was found
		for x in columns2.keys():
			if columns2[x] == []:
				del columns2[x]
		#Pass columns into a dataframe
		df3 = pd.DataFrame(columns2, index =[0])
		print "Did you know %s CIK is %s !" % (Name, i) 
		results2_df = results2_df.append(df3)
		try:	
		#Here we are trying to scrape the Filing infomration and links for each filing.  
			for file in Filing_list:
				type = '&type=' + file 
				footer = '&dateb=&owner=exclude&count=100'
				#Here we create a new url for the main sec page, with the filter for the type of filing we are currently looking for.
				url = header + str(i) + type + footer
				r = s.get(url)
				soup = bs(r.content,'html.parser')
				filing = soup.find(text=re.compile(file))
				table = soup.find_all('table')
				# file2 = file.encode('utf')
				# print filing
				# print file2
				# #if the file type exists somewhere in the html tree (after we fill for these types of filings) then we will scrape the page. 
				# m = re.search(file2, filing)
				if len(table) == 3:
					print file
					table = soup.find_all('table')
					Filings = []
					Format = []	
					Description	= [] 
					Filing_Date = []	
					File_Number = []
					CIK = []
					Link11 = []
					Link22 = []
					size11 = []
					size22 = []
					
					for row in table[2].find_all('tr')[1:]:
						col = row.find_all('td')
						col1 = col[0].string.strip()
						x =col1.encode('ascii','ignore')
						Filings.append(x)
						x = col[1].find('a').get('href')
						sec = 'https://www.sec.gov/'
						url2 = sec + x
						www = s.get(url2)
						soup2 = bs(www.content, 'html.parser')
						table2 = soup2.find_all('table')
						for num in (1, 3):
							for row2 in table2[0].find_all('tr')[1:]:
								try:
									col2 = row2.find_all('td')
									coll = col2[num].string.strip()
									coll = coll.encode('ascii','ignore')
									if re.search('Complete submission', coll):
											link2 = col2[2].find('a').get('href')
											link2 = link2.encode('ascii','ignore')
											link2 = sec + link2
											size2 = col2[4].string.strip()
											Link22.append(link2)
											size22.append(size2)
											continue
									elif re.search(file, coll):
											link1 = col2[2].find('a').get('href')
											link1 = link1.encode('ascii','ignore')
											link1 = sec + link1
											size1 = col2[4].string.strip()
											if len(Link11) == len(Link22):
												continue
											
											Link11.append(link1)
											size22.append(size1)
											pass
									elif len(Link11) < len(Link22):
										Link11.append('None')
											
																
								except: 
									pass
									
						col2 = col[3].string.strip()
						y =col2.encode('ascii','ignore')
						Filing_Date.append(y)
						
				
					columns = {'Filings' :Filings, 'Filing_Date' : Filing_Date, 'CIK' : i, "Company Name" : Name, "Filing Link" : Link11, "SIZE_Filing":size11,"Complete Submission Text" : Link22,"SIZE Complete Text": size22}
							
					df2 = pd.DataFrame(columns)
					
					results_df2 = results_df2.append(df2)
				else:
					pass
					
			
			progress = round((float(time2) / len(cik)) * 100)
			print "\n %s Did it" %(i)
			print "\n %d Percent Done" % (progress)
			if time2 >= 500 * loop:
				print "Saving Results"
				loop = 1 + loop
				savingout(result_company_info, results2_df)
				savingout(result_filing, results_df2)
				savingout(result_wrong, wrong_df)
				wrong_df = pd.DataFrame()
				results_df2 = pd.DataFrame()
				results2_df = pd.DataFrame()
				continue
			else:
				continue
			
		except:
			wrong = "No Description"
			columns = {'Wrong' : wrong, 'CIK': i}
			df = pd.DataFrame(columns, index =[0])
			wrong_df = wrong_df.append(df)
			print "Whoomps there it is %s (No Filings)" %(i) 
			progress = round((float(time2) / len(cik)) * 100)
			print "\n %d Percent Done" % (progress)
			if time2 >= 500 * loop:
				print "Saving Results"
				loop = 1 + loop
				savingout(result_company_info, results2_df)
				savingout(result_filing, results_df2)
				savingout(result_wrong, wrong_df)
				wrong_df = pd.DataFrame()
				results_df2 = pd.DataFrame()
				results2_df = pd.DataFrame()
				continue
			else:
				continue
			
			
	except:
		wrong = "No Filings"
		columns = {'Wrong' : wrong, 'CIK': i}
		df = pd.DataFrame(columns, index =[0])
		wrong_df = wrong_df.append(df)
		print "\n Whoomps there it is %s (No Name)" %(i) 
		progress = round((float(time2) / len(cik)) * 100)
		print "\n %d Percent Done" % (progress)
		continue
		
savingout(result_company_info, results2_df)
savingout(result_filing, results_df2)
savingout(result_wrong, wrong_df)

wrong_df = pd.DataFrame()

results_df2 = pd.DataFrame()
results2_df = pd.DataFrame()

try:
	wrong_df = pd.read_csv(result_wrong)
except:
	pass
	
results_df2 = pd.read_csv(result_filing)
results2_df = pd.read_csv(result_company_info)
	
writer = ExcelWriter(results)
in_file.to_excel(writer,'cik', index=False)	
results2_df.to_excel(writer, 'Company_Info',index=False ) 
results_df2.to_excel(writer, 'filings',index=False ) 

try:
	wrong_df.to_excel(writer,'no_cik',index=False )
except:
	pass

writer.save()

##Doing cool things in the end.
import openpyxl as op
try:
	from openpyxl.cell import get_column_letter, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string


wb = op.load_workbook(results)
sheet = wb.get_sheet_by_name('Company_Info')

 
for i in range ( 2, sheet.max_row+1):
    xx = 'D' +str(i)
    w = sheet['I' + str(i)] = '=HYPERLINK(%s, "click here")' %(xx)

sheet = wb.get_sheet_by_name('filings')

for i in range ( 2, sheet.max_row+1):
	x = 'C' +str(i)
	xx = 'D' +str(i)
	sheet['G' + str(i)] = '=HYPERLINK(%s, "Complete Submission Text")' %(x)
	sheet['H' + str(i)] = '=HYPERLINK(%s, "Filing Link")' %(xx)

wb.save(results)	

x = len(cik)

print "\n For %d CIKs it took"  % (x), time.time() - start_time, "seconds to find the filings"